# -*- coding: utf-8 -*-
"""
Created on Tue Aug 23 18:54:47 2016
@author: Boris Knyazev

This code is an attempt to reproduce results similar to
[1] Xie, D., Liang, L., Jin, L., Xu, J., & Li, M. (2015, October). 
    SCUT-FBP: A Benchmark Dataset for Facial Beauty Perception. 
    In Systems, Man, and Cybernetics (SMC), 2015 IEEE International Conference on (pp. 1821-1826). IEEE.

Disclaimer:
Measuring attractiveness or beauty is highly subjective, so this is just an experiment
In this dataset, I actually find many faces with low rating quite cute

"""

from __future__ import print_function

import numpy as np
import matplotlib.pyplot as plt
import os
import theano

from openpyxl import load_workbook
from PIL import Image
from sklearn.decomposition import PCA
from sklearn import svm
from sklearn.metrics import mean_absolute_error
from timeit import default_timer as timer

from theano import tensor as T
from theano.tensor.nnet import conv2d
from theano.tensor.signal import pool
from theano.tensor.nnet import relu

start = timer()

# options
SVR_kernel='rbf' 
PCA_dim = 50 # the number of PCA components (0 - no PCA)
color_Gabor = False # true to create colored Gabors (randomly)

# load data
# Dataset can be downloaded from http://www.hcii-lab.net/data/SCUT-FBP/EN/download.html
data_dir = '/home/boris/Project/data/images/SCUT-FBP/';
print('loading data')
wb = load_workbook(filename=data_dir+'Rating_Collection/Attractiveness label.xlsx', read_only=True)
ws = wb['Sheet1']
image_files = os.listdir(data_dir+'Data_Collection/')
image_list = []
labels=[]
# we want to resize images, because it is not reasonable to work with very large images
face_sz = (224,294)
for r in range(ws.min_row+1, ws.max_row+1):
  labels.append((ws.cell('B%d' % r).value))
  image_path = data_dir+'Data_Collection' + '/SCUT-FBP-%d.jpg' % int(str(r-1))
  img = Image.open(open(image_path)).resize(face_sz)
  img = np.asarray(img, dtype='float64')[35:259,:,:] / 256. # crop [40:296,:,:]
  print(image_path + ', size: ' + str(img.shape) + ', attractiveness: %1.2f' % labels[r-2])
  image_list.append(img)

labels = np.asarray(labels)
print('%d images and %d labels read' % (len(image_list), len(labels)))

# plot distribution of ratings
hist, bin_edges = np.histogram(labels, density=True, bins=50)
fig = plt.figure()
ax = fig.add_subplot(111)
plt.scatter(bin_edges[0:len(hist)], hist)
ax.set_xlabel('Attractiveness rating')
ax.set_ylabel('Frequency')
ax.set_title('Distribution of ratings')
plt.show()

# show less, average and more beautiful faces (just to check if data makes sense)
#least_beautiful = np.argmin(labels)
#avg_beautiful = np.argmin(abs(labels-np.mean(labels)))
#most_beautiful = np.argmax(labels)
less_beautiful = np.random.permutation(np.nonzero(labels<2)[0])[0]
avg_beautiful = np.random.permutation(np.nonzero((labels>2) & (labels<3))[0])[0]
more_beautiful = np.random.permutation(np.nonzero(labels>4.5)[0])[0]
fig = plt.figure()
ax = fig.add_subplot(111)
ax.set_title('Attractiveness (low): %1.2f' % labels[less_beautiful])
plt.imshow(image_list[less_beautiful])
fig = plt.figure()
ax = fig.add_subplot(111)
ax.set_title('Attractiveness (avg): %1.2f' % labels[avg_beautiful])
plt.imshow(image_list[avg_beautiful])
fig = plt.figure()
ax = fig.add_subplot(111)
ax.set_title('Attractiveness (high): %1.2f' % labels[more_beautiful])
plt.imshow(image_list[more_beautiful])
plt.show()

print('computing features using Gabors filters')

# Functions to generate a Gabor filter instead of skimage.filters.gabor_kernel
def gabor(params, sz):
    s = (sz[0]-1)/2
    x0, y0 = 0., 0.
    x,y = np.mgrid[-s:s+1,-s:s+1]
    x_modul, y_modul = transform_axis(x, y, params['theta'])
    x_gaus, y_gaus = transform_axis(x, y, params['beta'])
    scales_a = 1./(2*np.sqrt(np.pi)*params['stdx'])
    scales_b = 1./(2*np.sqrt(np.pi)*params['stdy'])
    gaus = np.exp(-np.pi*((x_gaus-x0)**2*scales_a**2 + (y_gaus-y0)**2*scales_b**2))
    w0 = 1./params['lambda']
    modul = np.exp(-1j*(2.*np.pi*w0*(x_modul-x0)+params['phi']))
    return gaus.reshape(sz)*modul.reshape(sz)
    
def transform_axis(x, y, theta):
    tx, ty = 0., 0. # x,y-translations
    sx, sy = 1., 1. # x,y-scaling
    T_transl = np.array([[1.,0.,tx], [0.,1.,ty], [0.,0.,1.]])
    T_rot1 = np.array([[1.,np.tan(theta/2),0.], [0.,1.,0.], [0.,0.,1.]])
    T_rot2 = np.array([[1.,0.,0.], [-np.sin(theta),1.,0.], [0.,0.,1.]])
    T_rot3 = T_rot1;
    pts = np.stack((x.flatten(), y.flatten(), np.ones((x.size))),axis=1).T;
    pts = T_transl.dot(T_rot3.dot(T_rot2.dot(T_rot1.dot(pts))));
    X = (pts[0,:]*sx).reshape(x.shape);
    Y = (pts[1,:]*sy).reshape(x.shape);
    return X,Y

#Generate Gabor filters
n_angles, n_scales, gammas = (6, 3, (0.4, 1.5)) if color_Gabor else (4, 3, range(1,2))
scales = np.linspace(0.3*np.pi,np.pi,n_scales)
thetas = np.linspace(0,np.pi,n_angles+1)[0:n_angles]
filter_size = (9,9)
kernels = []
for scale_x in scales:
    for theta in thetas:
        for sig_L in (0.3, 0.6):
            for gamma in gammas:
                kernels.append(gabor({'theta':theta,'beta':theta,'stdx':scale_x,'stdy':scale_x/gamma,
                'lambda':scale_x/sig_L,'phi':0.}, filter_size))

# Visualization of filters
# http://sklearn-theano.github.io/auto_examples/plot_overfeat_layer1_filters.html
def make_visual(layer_weights):
    max_scale = layer_weights.max(axis=-1).max(axis=-1)[...,
                                                        np.newaxis, np.newaxis]
    min_scale = layer_weights.min(axis=-1).min(axis=-1)[...,
                                                        np.newaxis, np.newaxis]
    return (255 * (layer_weights - min_scale) /
            (max_scale - min_scale)).astype('uint8')
                
def make_mosaic(layer_weights, sz):
    # Dirty hack (TM)
    lw_shape = layer_weights.shape
    lw = make_visual(layer_weights).reshape(sz[0], sz[1], *lw_shape[1:])
    lw = lw.transpose(0, 3, 1, 4, 2)
    lw = lw.reshape(sz[0] * lw_shape[-1], sz[1] * lw_shape[-2], lw_shape[1])
    return lw


def plot_filters(layer_weights, sz=(8,9), title=None, show=False):
    mosaic = make_mosaic(layer_weights, sz)
    plt.imshow(mosaic, interpolation='nearest')
    ax = plt.gca()
    ax.xaxis.set_visible(False)
    ax.yaxis.set_visible(False)
    if title is not None:
        plt.title(title)
    if show:
        plt.show()

# normalize images
for i, img in enumerate(image_list):
    image_list[i] = (img-img.mean())#/img.std() 

# instantiate 4D tensor for input
# from http://deeplearning.net/tutorial/lenet.html
input = T.tensor4(name='input')

if color_Gabor:
#    Make colored Gabor filters by (random) stacking of gray filters
    kernels = np.stack(kernels)
#    kernels = kernels[np.random.permutation(kernels.shape[0]),:,:]
    W_real = np.real(kernels.reshape(len(kernels)/3,3,filter_size[0],filter_size[1]))
else:
    W_real = np.real(np.tile(np.stack(kernels).reshape(len(kernels),filter_size[0],filter_size[1],1), (1,1,1,3)).transpose(0,3,1,2))

# normalize filters
for i in range(W_real.shape[0]):
    v = W_real[i,:,:,:]
    v = (v-v.mean())/v.std()*0.02
    v = v.flatten()
    v[v>0.05] = 0.05 # colored filters sometimes have too large values 
    W_real[i,:,:,:] = v.reshape(W_real[i,:,:,:].shape)
    
plot_filters(W_real, sz=(3,8), title='Gabor filters (Re)', show=True)
W = theano.shared(np.asarray(W_real, dtype='float32'), name ='W')

# convolution, max-pooling and ReLU
f1 = theano.function([input], conv2d(input, W))
f2 = theano.function([input], pool.pool_2d(input, (16, 16), ignore_border=True))
feat_maps = relu(f2(f1(np.asarray(image_list, dtype='float32').transpose(0,3,1,2))))
feat_maps = feat_maps.reshape(feat_maps.shape[0],np.prod(feat_maps.shape[1:]))

if PCA_dim > 0:
    pca = PCA(n_components=PCA_dim) 

# 10-fold cross-validation
print('performing cross-validation using SVM regression')
n_folds = 10

# shuffle samples
ids = np.random.permutation(feat_maps.shape[0])
feat_maps = feat_maps[ids,:]
labels_ = labels[ids]
n = len(labels_)/n_folds

PC = []
MAE = []
for fold_id in range(n_folds):
    test_ids = np.arange(fold_id*n,(fold_id+1)*n)
    train_ids = np.concatenate((np.arange(0,fold_id*n),
    np.arange((fold_id+1)*n,len(labels_))))
    assert(len(train_ids) == n*(n_folds-1))
    assert(len(test_ids) == n)
    feat_maps_train = feat_maps[train_ids,:]
    feat_maps_test = feat_maps[test_ids,:]
    if PCA_dim > 0:
        feat_maps_train = pca.fit_transform(feat_maps_train)
        feat_maps_test = pca.transform(feat_maps_test)
    
#    feature scaling
    feat_maps_train = (feat_maps_train - np.mean(feat_maps_train,0))#/np.std(feat_maps_train,0)
    feat_maps_test = (feat_maps_test - np.mean(feat_maps_test,0))#/np.std(feat_maps_test,0)

    clf = svm.SVR(kernel=SVR_kernel, C=1)
    clf.fit(feat_maps_train, labels_[train_ids])
    pred = clf.predict(feat_maps_test)
    PC.append(np.corrcoef(labels_[test_ids],pred)[0,1])
    MAE.append(mean_absolute_error(labels_[test_ids],pred))
    
    fig = plt.figure()
    plt.plot(np.arange(n),labels_[test_ids], label='true')
    plt.plot(np.arange(n),pred, label='predicted')
    plt.legend(loc='upper left')
    plt.title('fold %d, PC = %1.2f, MAE = %1.2f' % 
    (fold_id,np.corrcoef(labels_[test_ids],pred)[0,1],mean_absolute_error(labels_[test_ids],pred)))
    plt.show()

print('PC (Pearson correlation) mean = %1.2f (random guess = %1.2f)' % (np.mean(PC),np.corrcoef(labels_[test_ids],2.5*np.ones((len(test_ids),))+0.1*np.random.rand(len(test_ids),)
)[0,1]))
print('MAE (Mean absolute error) mean = %1.2f (random guess = %1.2f)' % (np.mean(MAE),mean_absolute_error(labels_[test_ids],2.5*np.ones((len(test_ids),)))))

end = timer()
print('Test took %1.2f sec' % (end - start))